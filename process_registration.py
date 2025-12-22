"""
Process Community Registration Excel file and create communities and caretakers via GraphQL
"""

import os
import sys
import time
import getpass
import configparser
import openpyxl
from gql import gql, Client
from gql.transport.requests import RequestsHTTPTransport
from requests_aws4auth import AWS4Auth
import boto3
from botocore.exceptions import ClientError
from typing import Dict, List, Optional, Tuple

# Global variable to store selected environment config
_env_config = {}


def load_environment_config(env_name: str, config_file: str = 'env.local') -> Dict[str, str]:
    """
    Load environment configuration from INI-style config file
    
    Args:
        env_name: Environment name ('DEV' or 'PRD')
        config_file: Path to the configuration file
    
    Returns:
        Dictionary with configuration values
    
    Raises:
        ValueError: If environment section not found
    """
    global _env_config
    
    config = configparser.ConfigParser()
    config.read(config_file)
    
    env_name_upper = env_name.upper()
    
    if env_name_upper not in config.sections():
        available = ', '.join(config.sections()) if config.sections() else 'None'
        raise ValueError(f"Environment '{env_name_upper}' not found in {config_file}. Available: {available}")
    
    _env_config = dict(config[env_name_upper])
    
    # Also set as environment variables for compatibility
    for key, value in _env_config.items():
        os.environ[key.upper()] = value
    
    return _env_config


def get_config(key: str, default: str = None) -> Optional[str]:
    """
    Get configuration value from loaded environment config
    
    Args:
        key: Configuration key (case-insensitive)
        default: Default value if key not found
    
    Returns:
        Configuration value or default
    """
    # Try from loaded config first
    value = _env_config.get(key.lower())
    if value:
        return value
    
    # Fall back to environment variable
    return os.getenv(key.upper(), default)


def print_graphql_for_console(operation_name: str, query_string: str, variables: Dict) -> None:
    """
    Print GraphQL query/mutation in a format that can be copied to AppSync Console
    
    Args:
        operation_name: Name of the operation (for display)
        query_string: The GraphQL query/mutation string
        variables: The variables dictionary
    """
    import json
    
    print("\n" + "─"*70)
    print(f"📋 GRAPHQL DEBUG - {operation_name}")
    print("─"*70)
    print("\n🔹 QUERY/MUTATION:")
    print("─"*50)
    print(query_string.strip())
    print("─"*50)
    print("\n🔹 VARIABLES:")
    print("─"*50)
    print(json.dumps(variables, indent=2, default=str))
    print("─"*50 + "\n")


def create_appsync_client(api_url: str = None, region: str = None, jwt_token: str = None, api_key: str = None, verbose: bool = False, use_bearer_prefix: bool = False):
    """
    Create an authenticated GraphQL client for AWS AppSync
    
    Supports multiple authentication methods (in priority order):
    1. Cognito JWT token authentication - if jwt_token is provided
    2. API Key authentication - if APPSYNC_API_KEY is provided
    3. IAM authentication (default) - uses AWS credentials
    
    Args:
        api_url: Your AWS AppSync GraphQL endpoint URL (defaults to env variable)
        region: AWS region (defaults to env variable or 'us-east-1')
        jwt_token: Optional JWT token from Cognito User Pool authentication
        api_key: Optional API key for AppSync API Key authentication
        verbose: If True, print detailed debugging information
    
    Returns:
        GQL Client instance
    """
    # Get configuration from environment variables if not provided
    if api_url is None:
        api_url = get_config('APPSYNC_API_URL')
        if not api_url:
            raise ValueError("APPSYNC_API_URL must be set in env.local or passed as parameter")
    
    if region is None:
        region = get_config('AWS_REGION', 'us-east-1')
    
    if verbose:
        print(f"\n  [DEBUG] AppSync Client Configuration:")
        print(f"    API URL: {api_url}")
        print(f"    Region: {region}")
        print(f"    JWT Token provided: {'Yes' if jwt_token else 'No'}")
        print(f"    API Key provided: {'Yes' if api_key else 'No'}")
        print(f"    APPSYNC_API_KEY env: {'Set' if get_config('APPSYNC_API_KEY') else 'Not set'}")
        print(f"    AWS_PROFILE env: {get_config('AWS_PROFILE') or 'Not set (using default)'}")
    
    # Priority 1: Use JWT token authentication (Cognito User Pool)
    if jwt_token:
        # Determine auth header value
        auth_header_value = f"Bearer {jwt_token}" if use_bearer_prefix else jwt_token
        
        if verbose:
            print(f"    Auth method: JWT Token (Cognito)")
            print(f"    Using Bearer prefix: {'Yes' if use_bearer_prefix else 'No'}")
            print(f"    Token preview: {jwt_token[:50]}...")
            # Decode and show JWT claims (without verification) for debugging
            try:
                import base64
                import json
                # JWT is base64 encoded: header.payload.signature
                payload_b64 = jwt_token.split('.')[1]
                # Add padding if needed
                payload_b64 += '=' * (4 - len(payload_b64) % 4)
                payload = json.loads(base64.urlsafe_b64decode(payload_b64))
                print(f"    JWT Claims:")
                print(f"      - sub (user ID): {payload.get('sub', 'N/A')}")
                print(f"      - email: {payload.get('email', 'N/A')}")
                print(f"      - cognito:groups: {payload.get('cognito:groups', 'None')}")
                print(f"      - token_use: {payload.get('token_use', 'N/A')}")
                print(f"      - iss (issuer): {payload.get('iss', 'N/A')}")
                print(f"      - exp: {payload.get('exp', 'N/A')}")
            except Exception as e:
                print(f"    Could not decode JWT: {e}")
        
        headers = {
            'Authorization': auth_header_value
        }
        transport = RequestsHTTPTransport(
            url=api_url,
            headers=headers,
            use_json=True,
        )
    # Priority 2: Check for API Key authentication
    elif api_key or get_config('APPSYNC_API_KEY'):
        api_key = api_key or get_config('APPSYNC_API_KEY')
        if verbose:
            print(f"    Auth method: API Key")
        headers = {
            'x-api-key': api_key
        }
        transport = RequestsHTTPTransport(
            url=api_url,
            headers=headers,
            use_json=True,
        )
    # Priority 3: Use IAM authentication (AWS Signature Version 4)
    else:
        aws_profile = get_config('AWS_PROFILE')
        
        # Get AWS credentials from ~/.aws/credentials
        session_kwargs = {}
        if aws_profile:
            session_kwargs['profile_name'] = aws_profile
        
        session = boto3.Session(**session_kwargs)
        credentials = session.get_credentials()
        
        if not credentials:
            raise ValueError(
                "AWS credentials not found in ~/.aws/credentials. "
                "Please configure AWS credentials using 'aws configure' or set AWS_PROFILE in env.local."
            )
        
        # Get the frozen credentials to access actual values
        frozen_credentials = credentials.get_frozen_credentials()
        
        if verbose:
            print(f"    Auth method: IAM (AWS Signature V4)")
            print(f"    Credentials source: ~/.aws/credentials")
            print(f"    AWS Profile: {aws_profile or 'default'}")
            print(f"    Access Key ID: {frozen_credentials.access_key[:8]}..." if frozen_credentials.access_key else "    Access Key ID: None")
            print(f"    Has Session Token: {'Yes' if frozen_credentials.token else 'No'}")
            print(f"    Region: {region}")
        
        auth = AWS4Auth(
            frozen_credentials.access_key,
            frozen_credentials.secret_key,
            region,
            'appsync',
            session_token=frozen_credentials.token,
        )
        
        # Create transport with AWS authentication
        transport = RequestsHTTPTransport(
            url=api_url,
            auth=auth,
            use_json=True,
        )
    
    # Create GraphQL client
    # Disable schema fetching since we have the schema defined in code
    # Some AppSync APIs don't support full introspection queries
    client = Client(
        transport=transport,
        fetch_schema_from_transport=False,
    )
    
    if verbose:
        print(f"    ✓ AppSync client created successfully\n")
    
    return client


def create_cognito_client(region: str = None):
    """
    Create a Cognito Identity Provider client
    
    Args:
        region: AWS region (defaults to env variable or 'us-east-1')
    
    Returns:
        boto3 Cognito Identity Provider client
    """
    if region is None:
        region = get_config('AWS_REGION', 'us-east-1')
    
    # Get AWS profile if specified
    aws_profile = get_config('AWS_PROFILE')
    
    # Create session with profile if specified
    session_kwargs = {}
    if aws_profile:
        session_kwargs['profile_name'] = aws_profile
    
    session = boto3.Session(**session_kwargs)
    cognito_client = session.client('cognito-idp', region_name=region)
    
    return cognito_client


def authenticate_cognito_user(user_pool_id: str, client_id: str, username: str, password: str, region: str = None) -> str:
    """
    Authenticate a user with Cognito User Pool and get JWT token
    
    Args:
        user_pool_id: Cognito User Pool ID
        client_id: Cognito User Pool App Client ID
        username: Username (email or username)
        password: User password
        region: AWS region (defaults to env variable or 'us-east-1')
    
    Returns:
        JWT ID token from User Pool authentication
    
    Raises:
        Exception: If authentication fails
    """
    if region is None:
        region = get_config('AWS_REGION', 'us-east-1')
    
    cognito_idp_client = create_cognito_client(region)
    
    try:
        # Authenticate with User Pool using the specified Client ID
        print(f"  Authenticating with App Client: {client_id}")
        print(f"  Username: {username}")
        
        auth_params = {
            'USERNAME': username,
            'PASSWORD': password,
        }
        
        response = cognito_idp_client.initiate_auth(
            ClientId=client_id,
            AuthFlow='USER_PASSWORD_AUTH',
            AuthParameters=auth_params
        )
        
        # Check if challenge is required
        if 'ChallengeName' in response:
            challenge_name = response['ChallengeName']
            if challenge_name == 'NEW_PASSWORD_REQUIRED':
                raise Exception("New password required. Please change your password first using the Cognito console or mobile app.")
            else:
                raise Exception(f"Authentication challenge required: {challenge_name}. Please complete the challenge first.")
        
        # Get the ID token (JWT)
        if 'AuthenticationResult' not in response:
            raise Exception("Authentication response missing AuthenticationResult")
        
        id_token = response['AuthenticationResult']['IdToken']
        print(f"  ✓ Successfully obtained JWT token")
        return id_token
        
    except ClientError as e:
        error_code = e.response.get('Error', {}).get('Code', '')
        error_message = e.response.get('Error', {}).get('Message', '')
        
        print(f"  ✗ Cognito API Error: {error_code}")
        print(f"  Error details: {error_message}")
        
        if error_code == 'NotAuthorizedException':
            raise Exception(f"Invalid username or password. Error: {error_message}")
        elif error_code == 'UserNotConfirmedException':
            raise Exception("User account is not confirmed. Please verify your email address first.")
        elif error_code == 'UserNotFoundException':
            raise Exception(f"User '{username}' not found in Cognito User Pool.")
        elif error_code == 'InvalidParameterException':
            if 'USER_PASSWORD_AUTH' in str(error_message):
                raise Exception(f"USER_PASSWORD_AUTH is not enabled for this App Client. Please enable it in Cognito User Pool settings (App clients → {client_id} → Authentication flows).")
            else:
                raise Exception(f"Invalid parameter: {error_message}. Please check your Cognito configuration.")
        elif error_code == 'ResourceNotFoundException':
            raise Exception(f"Resource not found: {error_message}. Please verify your COGNITO_USER_POOL_ID and COGNITO_CLIENT_ID are correct.")
        else:
            raise Exception(f"Authentication failed with error code '{error_code}': {error_message}")
            
    except Exception as e:
        # Re-raise if it's already a formatted exception
        if "Invalid username or password" in str(e) or "not confirmed" in str(e) or "challenge required" in str(e):
            raise
        raise Exception(f"Authentication error: {str(e)}")


def get_or_create_cognito_group(cognito_client, user_pool_id: str, community_id: str, community_name: str) -> str:
    """
    Get or create a Cognito group for a community
    
    Args:
        cognito_client: boto3 Cognito client
        user_pool_id: Cognito User Pool ID
        community_id: Community ID
        community_name: Community name
    
    Returns:
        Group name
    """
    # Create group name from community ID (sanitize for Cognito group name requirements)
    # Cognito group names can contain letters, numbers, spaces, and these characters: + = . , - @ _
    group_name = f"community-{community_id}"
    
    try:
        # Try to get the group
        cognito_client.get_group(
            GroupName=group_name,
            UserPoolId=user_pool_id
        )
        print(f"  ✓ Cognito group '{group_name}' already exists")
    except ClientError as e:
        error_code = e.response.get('Error', {}).get('Code', '')
        if error_code == 'ResourceNotFoundException':
            # Group doesn't exist, create it
            try:
                cognito_client.create_group(
                    GroupName=group_name,
                    UserPoolId=user_pool_id,
                    Description=f"Group for community: {community_name} (ID: {community_id})"
                )
                print(f"  ✓ Created Cognito group '{group_name}' for community '{community_name}'")
            except Exception as create_error:
                print(f"  ⚠ Warning: Could not create Cognito group '{group_name}': {str(create_error)}")
                raise
        else:
            print(f"  ⚠ Warning: Error checking Cognito group '{group_name}': {str(e)}")
            raise
    except Exception as e:
        print(f"  ⚠ Warning: Error checking Cognito group '{group_name}': {str(e)}")
        raise
    
    return group_name


def add_user_to_cognito(cognito_client, user_pool_id: str, email: str, first_name: str, last_name: str, group_name: str) -> bool:
    """
    Add a user to Cognito User Pool and assign to a group
    Cognito will automatically send temporary password and invitation email
    Email verification is left to the user (not auto-verified)
    
    Args:
        cognito_client: boto3 Cognito client
        user_pool_id: Cognito User Pool ID
        email: User email address (used as username)
        first_name: User first name
        last_name: User last name
        group_name: Cognito group name to assign user to
    
    Returns:
        True if successful, False otherwise
    """
    try:
        # Create user in Cognito using email as username
        # MessageAction='WELCOME' sends invitation email with temp password
        # email_verified is set to 'false' so user must verify their email
        print(f"    Creating user in Cognito User Pool...")
        cognito_client.admin_create_user(
            UserPoolId=user_pool_id,
            Username=email,  # Email is used as username
            UserAttributes=[
                {'Name': 'email', 'Value': email},
                {'Name': 'email_verified', 'Value': 'false'},  # User must verify email
                {'Name': 'given_name', 'Value': first_name},
                {'Name': 'family_name', 'Value': last_name},
            ],
            #MessageAction='WELCOME'  # Send welcome message with temp password
        )
        print(f"    ✓ User created in Cognito")
        
        # Add user to group
        print(f"    Adding user to group '{group_name}'...")
        cognito_client.admin_add_user_to_group(
            UserPoolId=user_pool_id,
            Username=email,
            GroupName=group_name
        )
        print(f"    ✓ User added to group")
        
        print(f"  ✓ Added user to Cognito and assigned to group '{group_name}'")
        return True
        
    except ClientError as e:
        error_code = e.response.get('Error', {}).get('Code', '')
        error_message = e.response.get('Error', {}).get('Message', '')
        
        if error_code == 'UsernameExistsException':
            # User already exists, update attributes and add to group
            try:
                print(f"  User already exists in Cognito, updating and adding to group...")
                # Update user attributes (keep email_verified as false - user must verify)
                cognito_client.admin_update_user_attributes(
                    UserPoolId=user_pool_id,
                    Username=email,
                    UserAttributes=[
                        {'Name': 'email_verified', 'Value': 'false'},  # User must verify email
                        {'Name': 'given_name', 'Value': first_name},
                        {'Name': 'family_name', 'Value': last_name},
                    ]
                )
                
                # Add user to group
                cognito_client.admin_add_user_to_group(
                    UserPoolId=user_pool_id,
                    Username=email,
                    GroupName=group_name
                )
                
                print(f"  ✓ User already exists in Cognito, updated and added to group '{group_name}'")
                return True
            except ClientError as update_error:
                update_error_code = update_error.response.get('Error', {}).get('Code', '')
                update_error_message = update_error.response.get('Error', {}).get('Message', '')
                print(f"  ✗ Error updating existing user: {update_error_code} - {update_error_message}")
                return False
            except Exception as e:
                print(f"  ✗ Error updating existing user: {str(e)}")
                return False
        else:
            # Other ClientError - show detailed error
            print(f"  ✗ Cognito API Error: {error_code}")
            print(f"  Error message: {error_message}")
            return False
    except Exception as e:
        print(f"  ✗ Unexpected error adding user to Cognito: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def read_community_data(file_path: str) -> List[Dict]:
    """
    Read community data from Excel file
    
    Args:
        file_path: Path to the Excel file
    
    Returns:
        List of community data dictionaries
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Community Info']
    
    communities = []
    headers = [cell.value for cell in ws[1]]
    
    # Map Excel column names to GraphQL input field names
    field_mapping = {
        'Name': 'name',
        'Contact Phone Number': 'phoneNumber',
        'Contact Email': 'email',
        'Street': 'street',
        'City': 'city',
        'State': 'state',
        'Country': 'country',
        'Zip Code': 'postalCode',
        'No. Resident': 'residentLimit',
        'No. Users': 'caretakerLimit',
    }
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not any(row):
            continue
        
        community = {}
        for idx, header in enumerate(headers):
            if header in field_mapping and row[idx] is not None:
                field_name = field_mapping[header]
                value = row[idx]
                
                # Convert numeric fields to int
                if field_name in ['residentLimit', 'caretakerLimit']:
                    value = int(value)
                
                community[field_name] = value
        
        # Only add if required fields are present
        if community.get('name') and community.get('phoneNumber') and community.get('email'):
            # Set defaults for required fields if not provided
            if 'residentLimit' not in community:
                community['residentLimit'] = 100
            if 'caretakerLimit' not in community:
                community['caretakerLimit'] = 10
            
            communities.append(community)
    
    return communities


def read_caretaker_data(file_path: str) -> List[Dict]:
    """
    Read caretaker data from Excel file
    
    Args:
        file_path: Path to the Excel file
    
    Returns:
        List of caretaker data dictionaries
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Users']
    
    caretakers = []
    headers = [cell.value for cell in ws[1]]
    
    # Map Excel column names to GraphQL input field names
    field_mapping = {
        'First Name': 'firstName',
        'Last Name': 'lastName',
        'Email': 'email',
        'CommunityId': 'communityId',
    }
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not any(row):
            continue
        
        caretaker = {}
        for idx, header in enumerate(headers):
            if header in field_mapping and row[idx] is not None:
                field_name = field_mapping[header]
                caretaker[field_name] = row[idx]
        
        # Only add if required fields are present
        if (caretaker.get('firstName') and 
            caretaker.get('lastName') and 
            caretaker.get('email')):
            caretakers.append(caretaker)
    
    return caretakers


def update_excel_with_community_id(file_path: str, community_id: str) -> None:
    """
    Update the Excel file with the community ID in the Users sheet
    
    Args:
        file_path: Path to the Excel file
        community_id: The community ID to write
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb['Users']
        
        # Find the CommunityId column index
        headers = [cell.value for cell in ws[1]]
        community_id_col_idx = None
        
        for idx, header in enumerate(headers):
            if header == 'CommunityId':
                community_id_col_idx = idx + 1  # openpyxl uses 1-based indexing
                break
        
        # If CommunityId column doesn't exist, create it
        if community_id_col_idx is None:
            # Add CommunityId header at the end
            max_col = ws.max_column
            ws.cell(row=1, column=max_col + 1, value='CommunityId')
            community_id_col_idx = max_col + 1
        
        # Find First Name column index (to check if row has data)
        first_name_col = None
        for idx, header in enumerate(headers):
            if header == 'First Name':
                first_name_col = idx + 1
                break
        
        # Update all rows with the community ID (overwrite any existing values)
        for row_idx in range(2, ws.max_row + 1):
            # Only update rows that have data (check if first name exists)
            if first_name_col and ws.cell(row=row_idx, column=first_name_col).value:
                old_value = ws.cell(row=row_idx, column=community_id_col_idx).value
                ws.cell(row=row_idx, column=community_id_col_idx, value=community_id)
                if old_value and str(old_value) != str(community_id):
                    print(f"    Updated row {row_idx}: '{old_value}' → '{community_id}'")
        
        # Save the workbook
        wb.save(file_path)
        print(f"  ✓ Updated Excel file with CommunityId: {community_id}")
        
    except Exception as e:
        print(f"  ⚠ Warning: Could not update Excel file with CommunityId: {str(e)}")


def create_community(client: Client, community_data: Dict, verbose: bool = False) -> Optional[Dict]:
    """
    Create a community using GraphQL mutation
    
    Args:
        client: GraphQL client
        community_data: Community data dictionary
        verbose: If True, print detailed debugging information
    
    Returns:
        Created community data or None if failed
    """
    mutation_string = """
        mutation CreateCommunity($input: CreateCommunityInput!) {
            createCommunity(input: $input) {
                id
                name
                street
                city
                state
                country
                postalCode
                phoneNumber
                email
                residentLimit
                caretakerLimit
                isActive
                createdAt
                updatedAt
            }
        }
    """
    mutation = gql(mutation_string)
    variables = {'input': community_data}
    
    if verbose:
        print_graphql_for_console("CREATE COMMUNITY", mutation_string, variables)
    
    try:
        if verbose:
            print(f"  [VERBOSE] Executing mutation...")
        result = client.execute(mutation, variable_values=variables)
        
        if verbose:
            print(f"  [VERBOSE] Mutation result: {result}")
        
        return result['createCommunity']
    except Exception as e:
        error_details = str(e)
        if hasattr(e, 'errors') and e.errors:
            error_details = e.errors
        elif isinstance(e, dict):
            error_details = e
        
        print(f"Error creating community '{community_data.get('name')}': {error_details}")
        
        # Always print GraphQL debug on error so user can test in console
        print_graphql_for_console("CREATE COMMUNITY (FAILED)", mutation_string, variables)
        
        if verbose:
            print(f"\n  [VERBOSE] Error Details:")
            print(f"    Error type: {type(e).__name__}")
            print(f"    Error message: {str(e)}")
            if hasattr(e, 'errors'):
                print(f"    GraphQL errors: {e.errors}")
            if hasattr(e, 'response'):
                print(f"    Response: {e.response}")
            import traceback
            print(f"    Traceback:")
            traceback.print_exc()
        
        return None


def verify_caretaker_created(client: Client, email: str) -> bool:
    """
    Verify that a caretaker was created correctly by querying getUserByEmail
    
    Args:
        client: GraphQL client
        email: Email address of the caretaker to verify
    
    Returns:
        True if caretaker is found, False otherwise
    """
    query = gql("""
        query GetUserByEmail($email: String!, $role: String!) {
            getUserByEmail(email: $email, role: $role) {
                id
                email
                firstName
                lastName
                communityId
                role
                isActive
            }
        }
    """)
    
    try:
        result = client.execute(query, variable_values={
            'email': email,
            'role': 'CARETAKER'
        })
        users = result.get('getUserByEmail', [])
        return len(users) > 0
    except Exception as e:
        print(f"  ⚠ Verification query error: {str(e)}")
        return False


def create_caretaker(client: Client, caretaker_data: Dict, verbose: bool = False) -> Optional[Dict]:
    """
    Create a caretaker using GraphQL mutation
    
    Args:
        client: GraphQL client
        caretaker_data: Caretaker data dictionary
        verbose: If True, print detailed debugging information
    
    Returns:
        Created caretaker data or None if failed
    """
    mutation_string = """
        mutation CreateCommunityCaretaker($input: CreateCaretakerInput!) {
            createCommunityCaretaker(input: $input) {
                id
                communityId
                firstName
                lastName
                email
                role
                isActive
                createdAt
                updatedAt
            }
        }
    """
    mutation = gql(mutation_string)
    variables = {'input': caretaker_data}
    
    if verbose:
        print_graphql_for_console("CREATE CARETAKER", mutation_string, variables)
    
    try:
        if verbose:
            print(f"  [VERBOSE] Executing mutation...")
        result = client.execute(mutation, variable_values=variables)
        
        if verbose:
            print(f"  [VERBOSE] Mutation result: {result}")
        
        return result['createCommunityCaretaker']
    except Exception as e:
        error_details = str(e)
        if hasattr(e, 'errors') and e.errors:
            error_details = e.errors
        elif isinstance(e, dict):
            error_details = e
        
        print(f"Error creating caretaker '{caretaker_data.get('firstName')} {caretaker_data.get('lastName')}': {error_details}")
        
        # Always print GraphQL debug on error so user can test in console
        print_graphql_for_console("CREATE CARETAKER (FAILED)", mutation_string, variables)
        
        if verbose:
            print(f"\n  [VERBOSE] Error Details:")
            print(f"    Error type: {type(e).__name__}")
            print(f"    Error message: {str(e)}")
            if hasattr(e, 'errors'):
                print(f"    GraphQL errors: {e.errors}")
            if hasattr(e, 'response'):
                print(f"    Response: {e.response}")
            import traceback
            print(f"    Traceback:")
            traceback.print_exc()
        
        return None


def print_progress_header(phase: str, step: int, total_steps: int, description: str = ""):
    """
    Print a progress header showing overall progress
    
    Args:
        phase: Current phase name
        step: Current step number
        total_steps: Total number of steps
        description: Optional description
    """
    percentage = int((step / total_steps) * 100)
    progress_bar_length = 40
    filled = int((step / total_steps) * progress_bar_length)
    bar = "█" * filled + "░" * (progress_bar_length - filled)
    
    print("\n" + "="*60)
    print(f"OVERALL PROGRESS: [{bar}] {percentage}%")
    print(f"Phase {step}/{total_steps}: {phase}")
    if description:
        print(f"  {description}")
    print("="*60)


def process_excel_file(file_path: str, verbose: bool = False, use_iam: bool = False, use_bearer: bool = False) -> Dict:
    """
    Process the entire Excel file and create communities and caretakers
    
    Args:
        file_path: Path to the Excel file
        verbose: If True, print detailed debugging information
        use_iam: If True, use IAM authentication instead of Cognito JWT
        use_bearer: If True, add "Bearer" prefix to Authorization header
    
    Returns:
        Dictionary with summary of created records
    """
    # Define total steps for progress tracking
    # Step 1: Reading data, Step 2: Creating community, Step 3: Creating caretakers
    TOTAL_STEPS = 3
    
    # Cognito User Pool ID is always needed for user management
    cognito_user_pool_id = get_config('COGNITO_USER_POOL_ID')
    cognito_client_id = get_config('COGNITO_CLIENT_ID')
    
    if not cognito_user_pool_id:
        print("\n" + "="*60)
        print("ERROR: COGNITO_USER_POOL_ID is required")
        print("="*60)
        print("Cognito User Pool ID is required for user registration.")
        print("Please set COGNITO_USER_POOL_ID in your .env file.")
        sys.exit(1)
    
    # Create GraphQL client based on auth mode
    if use_iam:
        # Use IAM authentication (AWS credentials from ~/.aws/credentials)
        print("="*60)
        print("Using IAM Authentication")
        print("="*60)
        aws_profile = get_config('AWS_PROFILE')
        print(f"Credentials source: ~/.aws/credentials")
        print(f"AWS Profile: {aws_profile or 'default'}")
        
        # Verify credentials exist before proceeding
        session_kwargs = {}
        if aws_profile:
            session_kwargs['profile_name'] = aws_profile
        session = boto3.Session(**session_kwargs)
        credentials = session.get_credentials()
        if not credentials:
            print("\n" + "="*60)
            print("ERROR: AWS credentials not found")
            print("="*60)
            print("Could not find credentials in ~/.aws/credentials")
            print(f"Looking for profile: {aws_profile or 'default'}")
            print("\nTo configure AWS credentials, run:")
            print("  aws configure")
            print("\nOr specify a different profile in env.local:")
            print("  AWS_PROFILE=your-profile-name")
            sys.exit(1)
        
        frozen_creds = credentials.get_frozen_credentials()
        print(f"Access Key ID: {frozen_creds.access_key[:8]}...")
        print("="*60)
        
        # Create GraphQL client with IAM auth (no JWT token)
        client = create_appsync_client(verbose=verbose)
    else:
        # Use Cognito JWT authentication
        if not cognito_client_id:
            print("\n" + "="*60)
            print("ERROR: COGNITO_CLIENT_ID is required")
            print("="*60)
            print("Cognito App Client ID is required for JWT authentication.")
            print("Please set COGNITO_CLIENT_ID in your .env file.")
            print("Or use --iam flag to use IAM authentication instead.")
            sys.exit(1)
        
        # Prompt for username and password
        print("="*60)
        print("Cognito Authentication Required")
        print("="*60)
        print("Enter your credentials to authenticate with Cognito")
        print("(This is for GraphQL API access)")
        print("="*60)
        username = input("Enter your username (email): ").strip()
        if not username:
            print("ERROR: Username cannot be empty")
            sys.exit(1)
        
        password = getpass.getpass("Enter your password: ")
        if not password:
            print("ERROR: Password cannot be empty")
            sys.exit(1)
        
        print("\nAuthenticating with Cognito...")
        try:
            jwt_token = authenticate_cognito_user(
                cognito_user_pool_id,
                cognito_client_id,
                username,
                password
            )
            print("  ✓ Authentication successful")
        except Exception as e:
            print(f"  ✗ Authentication failed: {str(e)}")
            sys.exit(1)
        
        # Create GraphQL client with JWT token
        client = create_appsync_client(jwt_token=jwt_token, verbose=verbose, use_bearer_prefix=use_bearer)
    
    # Initialize Cognito client (REQUIRED)
    try:
        cognito_client = create_cognito_client()
        print("  ✓ Cognito client initialized")
    except Exception as e:
        print("\n" + "="*60)
        print("ERROR: Could not initialize Cognito client")
        print("="*60)
        print(f"Error: {str(e)}")
        print("Cognito is required for user authentication. Cannot proceed.")
        sys.exit(1)
    
    cognito_group_name = None
    
    # Step 1: Read data from Excel
    print_progress_header("Reading Excel File", 1, TOTAL_STEPS, "Extracting community and caretaker data...")
    communities = read_community_data(file_path)
    caretakers = read_caretaker_data(file_path)
    
    # Validate that there is exactly one community
    if len(communities) == 0:
        print("\n" + "="*60)
        print("ERROR: No valid community found in Excel file")
        print("="*60)
        print("The Excel file must contain exactly one community in the 'Community Info' sheet.")
        sys.exit(1)
    elif len(communities) > 1:
        print("\n" + "="*60)
        print("ERROR: Multiple communities found in Excel file")
        print("="*60)
        print(f"Found {len(communities)} communities:")
        for idx, comm in enumerate(communities, 1):
            print(f"  {idx}. {comm.get('name', 'Unknown')}")
        print("\nThe Excel file must contain exactly one community in the 'Community Info' sheet.")
        sys.exit(1)
    
    print(f"  ✓ Found 1 community and {len(caretakers)} caretakers to create")
    
    # Step 2: Create community (only one)
    print_progress_header("Creating Community", 2, TOTAL_STEPS, f"Creating community: {communities[0].get('name') if communities else 'Unknown'}")
    
    created_communities = []
    community_id = None
    
    community_data = communities[0]
    print(f"\nCreating community: {community_data.get('name')}")
    
    if verbose:
        print(f"\n  [VERBOSE] Community Data:")
        print(f"    {community_data}")
    
    result = create_community(client, community_data, verbose=verbose)
    
    if result:
        created_communities.append(result)
        community_id = result['id']
        print(f"  ✓ Successfully created with ID: {community_id}")
        
        if verbose:
            print(f"\n  [VERBOSE] Community Creation Result:")
            print(f"    Community ID: {community_id}")
            print(f"    Full result: {result}")
        
        # Update Excel file with community ID
        update_excel_with_community_id(file_path, community_id)
        
        # Create/get Cognito group for this community (REQUIRED)
        try:
            cognito_group_name = get_or_create_cognito_group(
                cognito_client,
                cognito_user_pool_id,
                community_id,
                community_data.get('name', 'Unknown')
            )
        except Exception as e:
            print(f"  ✗ Failed to create/get Cognito group")
            print("\n" + "="*60)
            print("ERROR: Cognito group creation failed")
            print("="*60)
            print(f"Error: {str(e)}")
            print("Cognito group is required for user authentication. Cannot proceed.")
            sys.exit(1)
        
        # Wait for DynamoDB to propagate changes
        print(f"\n  Waiting for DynamoDB to propagate changes...")
        wait_seconds = 3
        for i in range(wait_seconds, 0, -1):
            print(f"  Waiting {i} second(s)...", end='\r')
            time.sleep(1)
        print(f"  ✓ Ready to create caretakers")
    else:
        print(f"  ✗ Failed to create")
        print("\n" + "="*60)
        print("ERROR: Failed to create community. Cannot proceed with caretaker creation.")
        print("="*60)
        sys.exit(1)
    
    # Step 3: Create caretakers
    print_progress_header("Creating Caretakers", 3, TOTAL_STEPS, f"Creating {len(caretakers)} caretaker(s)...")
    
    created_caretakers = []
    
    for idx, caretaker_data in enumerate(caretakers, 1):
        # Show progress within caretaker creation phase
        caretaker_progress = int((idx / len(caretakers)) * 100) if len(caretakers) > 0 else 0
        print(f"\n[{idx}/{len(caretakers)}] ({caretaker_progress}%) Creating caretaker: {caretaker_data.get('firstName')} {caretaker_data.get('lastName')}")
        
        # Always use the newly created community ID (overwrite any existing value from Excel)
        old_community_id = caretaker_data.get('communityId')
        caretaker_data['communityId'] = community_id
        
        if old_community_id and old_community_id != community_id:
            print(f"  ⚠ Overwriting existing communityId '{old_community_id}' with newly created '{community_id}'")
        
        if verbose:
            print(f"\n  [VERBOSE] Before creating caretaker:")
            print(f"    Caretaker data (before): {caretaker_data}")
            if old_community_id and old_community_id != community_id:
                print(f"    ⚠ WARNING: Overwriting existing communityId '{old_community_id}' with newly created '{community_id}'")
            print(f"    Community ID from variable: {community_id}")
            print(f"    Community ID in caretaker_data (after update): {caretaker_data.get('communityId')}")
            print(f"    Types: community_id={type(community_id)}, caretaker_data['communityId']={type(caretaker_data.get('communityId'))}")
        
        result = create_caretaker(client, caretaker_data, verbose=verbose)
        
        if result:
            created_caretakers.append(result)
            print(f"  ✓ Successfully created with ID: {result['id']}")
            
            # Verify caretaker was created correctly (round-trip check)
            caretaker_email = caretaker_data.get('email')
            if not caretaker_email:
                print(f"  ✗ Cannot proceed: email not found in caretaker data")
                print("\n" + "="*60)
                print("ERROR: Email is required for verification and Cognito registration")
                print("="*60)
                sys.exit(1)
            
            print(f"  Verifying caretaker creation...")
            is_verified = verify_caretaker_created(client, caretaker_email)
            if is_verified:
                print(f"  ✓ Verification successful: Caretaker found in system")
            else:
                print(f"  ⚠ ALARM: Verification failed! Caretaker '{caretaker_data.get('firstName')} {caretaker_data.get('lastName')}' (email: {caretaker_email}) was not found after creation.")
                print(f"  ⚠ The caretaker may not have been created correctly. Please verify manually.")
            
            # Add user to Cognito and assign to group (REQUIRED)
            if not cognito_group_name:
                print(f"  ✗ Cannot add to Cognito: group name not set")
                print("\n" + "="*60)
                print("ERROR: Cognito group name is required")
                print("="*60)
                sys.exit(1)
            
            print(f"  Adding user to Cognito...")
            print(f"    User Pool ID: {cognito_user_pool_id}")
            print(f"    Group: {cognito_group_name}")
            first_name = caretaker_data.get('firstName', '')
            last_name = caretaker_data.get('lastName', '')
            
            try:
                cognito_success = add_user_to_cognito(
                    cognito_client,
                    cognito_user_pool_id,
                    caretaker_email,
                    first_name,
                    last_name,
                    cognito_group_name
                )
                
                if not cognito_success:
                    print(f"  ✗ Failed to add user to Cognito")
                    print("\n" + "="*60)
                    print("ERROR: Cognito user registration failed")
                    print("="*60)
                    print(f"User '{caretaker_data.get('firstName')} {caretaker_data.get('lastName')}' (email: {caretaker_email})")
                    print("was created in GraphQL but failed to register in Cognito.")
                    print("User authentication will not work. Cannot proceed.")
                    sys.exit(1)
            except Exception as e:
                print(f"  ✗ Exception while adding user to Cognito: {str(e)}")
                print("\n" + "="*60)
                print("ERROR: Cognito user registration failed")
                print("="*60)
                print(f"User '{caretaker_data.get('firstName')} {caretaker_data.get('lastName')}' (email: {caretaker_email})")
                print("was created in GraphQL but failed to register in Cognito.")
                print(f"Error: {str(e)}")
                print("User authentication will not work. Cannot proceed.")
                sys.exit(1)
        else:
            print(f"  ✗ Failed to create")
    
    # Show completion progress
    print("\n" + "="*60)
    print("OVERALL PROGRESS: [" + "█" * 40 + "] 100%")
    print("Phase 3/3: Creating Caretakers - COMPLETE")
    print("="*60)
    
    # Summary
    summary = {
        'communities': {
            'total': len(communities),
            'created': len(created_communities),
            'failed': len(communities) - len(created_communities),
            'data': created_communities
        },
        'caretakers': {
            'total': len(caretakers),
            'created': len(created_caretakers),
            'failed': len(caretakers) - len(created_caretakers),
            'data': created_caretakers
        }
    }
    
    return summary


def select_environment() -> str:
    """
    Prompt user to select environment (DEV or PRD)
    
    Returns:
        Selected environment name
    """
    print("\n" + "="*60)
    print("ENVIRONMENT SELECTION")
    print("="*60)
    print("\nAvailable environments:")
    print("  1. DEV  - Development environment")
    print("  2. PRD  - Production environment")
    print("")
    
    while True:
        choice = input("Select environment (1 for DEV, 2 for PRD): ").strip()
        
        if choice == '1':
            print("\n  ✓ Selected: DEV (Development)")
            return 'DEV'
        elif choice == '2':
            # Show production warning
            print("\n" + "!"*60)
            print("⚠️  WARNING: PRODUCTION ENVIRONMENT SELECTED ⚠️")
            print("!"*60)
            print("")
            print("You are about to work with the PRODUCTION environment.")
            print("This will create REAL communities and users in the live system.")
            print("")
            print("!"*60)
            
            confirm = input("\nAre you sure you want to continue with PRODUCTION? (type 'yes' to confirm): ").strip().lower()
            
            if confirm == 'yes':
                print("\n  ✓ Confirmed: PRD (Production)")
                return 'PRD'
            else:
                print("\n  ✗ Production not confirmed. Please select again.\n")
        else:
            print("  Invalid choice. Please enter 1 or 2.")


def main():
    """Main function to process registration"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Process Community Registration Excel file')
    parser.add_argument('file', help='Path to Excel file')
    parser.add_argument('--verbose', '-v', action='store_true',
                       help='Enable verbose output for debugging')
    parser.add_argument('--iam', action='store_true',
                       help='Use IAM authentication instead of Cognito JWT (use this if you have AWS credentials configured)')
    parser.add_argument('--bearer', action='store_true',
                       help='Add "Bearer" prefix to Authorization header (try this if getting Unauthorized errors)')
    parser.add_argument('--env', '-e', choices=['DEV', 'PRD', 'dev', 'prd'],
                       help='Environment to use (DEV or PRD). If not specified, will prompt for selection.')
    
    args = parser.parse_args()
    
    verbose = args.verbose
    use_iam = args.iam
    use_bearer = args.bearer
    
    if not os.path.exists(args.file):
        print(f"Error: File '{args.file}' not found")
        return
    
    # Select or confirm environment
    if args.env:
        env_name = args.env.upper()
        if env_name == 'PRD':
            # Show production warning even when specified via command line
            print("\n" + "!"*60)
            print("⚠️  WARNING: PRODUCTION ENVIRONMENT SELECTED ⚠️")
            print("!"*60)
            print("")
            print("You are about to work with the PRODUCTION environment.")
            print("This will create REAL communities and users in the live system.")
            print("")
            print("!"*60)
            
            confirm = input("\nAre you sure you want to continue with PRODUCTION? (type 'yes' to confirm): ").strip().lower()
            
            if confirm != 'yes':
                print("\n  ✗ Production not confirmed. Exiting.")
                sys.exit(0)
            print("\n  ✓ Confirmed: PRD (Production)")
        else:
            print(f"\n  ✓ Using environment: {env_name}")
    else:
        env_name = select_environment()
    
    # Load environment configuration
    try:
        load_environment_config(env_name)
        print(f"  ✓ Configuration loaded for {env_name}")
    except ValueError as e:
        print(f"\nError: {str(e)}")
        sys.exit(1)
    
    print("\n" + "="*60)
    print("Community Registration Processor")
    print("="*60)
    print(f"Environment: {env_name}")
    print(f"File: {args.file}")
    print(f"API URL: {get_config('APPSYNC_API_URL')}")
    print(f"Region: {get_config('AWS_REGION', 'us-east-1')}")
    print(f"Auth Mode: {'IAM' if use_iam else 'Cognito JWT'}")
    if use_bearer:
        print(f"Bearer Prefix: Enabled")
    print("="*60)
    
    try:
        summary = process_excel_file(args.file, verbose=verbose, use_iam=use_iam, use_bearer=use_bearer)
        
        # Print summary
        print("\n" + "="*60)
        print("SUMMARY")
        print("="*60)
        print(f"\nCommunities:")
        print(f"  Total: {summary['communities']['total']}")
        print(f"  Created: {summary['communities']['created']}")
        print(f"  Failed: {summary['communities']['failed']}")
        
        print(f"\nCaretakers:")
        print(f"  Total: {summary['caretakers']['total']}")
        print(f"  Created: {summary['caretakers']['created']}")
        print(f"  Failed: {summary['caretakers']['failed']}")
        
        print("\n" + "="*60)
        print("Processing complete!")
        print("="*60)
        
    except Exception as e:
        print(f"\nError processing file: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()

