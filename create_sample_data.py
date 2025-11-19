"""
Create a sample Excel file with test data for community registration
"""

import openpyxl
from openpyxl import Workbook

def create_sample_excel(output_file: str = 'sample_registration.xlsx'):
    """Create a sample Excel file with test data"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Community Info sheet
    ws_community = wb.create_sheet('Community Info')
    
    # Headers
    ws_community.append([
        'Name',
        'Contact Phone Number',
        'Contact Email',
        'Street',
        'City',
        'State',
        'Country',
        'Zip Code',
        'No. Resident',
        'No. Users',
        'CommunityId'
    ])
    
    # Sample data
    ws_community.append([
        'Sunrise Senior Living',
        '+1-555-0101',
        'contact@sunrisesenior.com',
        '123 Sunrise Boulevard',
        'San Francisco',
        'CA',
        'USA',
        '94102',
        150,
        15,
        None  # Will be filled after creation
    ])
    
    ws_community.append([
        'Golden Years Community',
        '+1-555-0202',
        'info@goldenyears.com',
        '456 Golden Lane',
        'Los Angeles',
        'CA',
        'USA',
        '90001',
        200,
        20,
        None
    ])
    
    # Create Users sheet
    ws_users = wb.create_sheet('Users')
    
    # Headers
    ws_users.append([
        'First Name',
        'Last Name',
        'Email',
        'CommunityId'
    ])
    
    # Sample data
    ws_users.append([
        'John',
        'Doe',
        'john.doe@sunrisesenior.com',
        None  # Will be filled after community creation
    ])
    
    ws_users.append([
        'Jane',
        'Smith',
        'jane.smith@sunrisesenior.com',
        None
    ])
    
    ws_users.append([
        'Michael',
        'Johnson',
        'michael.j@goldenyears.com',
        None
    ])
    
    # Save workbook
    wb.save(output_file)
    print(f"Sample Excel file created: {output_file}")
    print("\nThis file contains:")
    print("  - 2 sample communities in 'Community Info' sheet")
    print("  - 3 sample caretakers in 'Users' sheet")
    print("\nYou can edit this file with your actual data before processing.")


if __name__ == "__main__":
    create_sample_excel()

